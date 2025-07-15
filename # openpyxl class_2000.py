# openpyxl class_2000

from openpyxl import *
import os

class excel1():

    def __init__(self):
        pass

    def create(self):
        wb = Workbook()
        ws = wb.active

        name = str(input("\nplease enter the file's name : "))
        sheet_name = str(input("\nplease enter the sheet name : "))
        where  = str(input("\nwhere to save \nexample-->> C:\\Users\\PC\\Desktop \n : "))
        
        ws.title = sheet_name
        file_name=name+".xlsx"
        full_path = os.path.join(where, file_name)
        wb.save(full_path)
        print(f"\n''{name}''  was saved to: ''{where}''")
        
    def edit(self):
        file_path = str(input("\nplease enter the file's path and name\nexample-->> C:\\Users\\PC\\Desktop\\sample.xlsx \n : "))

        wb = load_workbook(file_path)
        ws = wb.active
        
        q = str(input("\n1.create new sheet 2.insert data \n : "))
        if q=="1":
            
            sheet_name = str(input("\nplease enter the sheet name : "))
            ws_data = wb.create_sheet(sheet_name)
            wb.save(file_path)
            print(f"\n''{sheet_name}''  was added to: ''{file_path}''")
            
        if q=="2":
            
            data = str(input("\nplease enter the data : "))
            
            sheet = wb.create_sheet(f'{str(input("\nplease enter the sheet name : "))}')
            
            cell =  str(input("\nplease enter the cell \nexample-->> B5\n: "))
        
            sheet[cell]=data
            
            wb.save(file_path)
            print(f"\n''{data}''  was added to: ''{file_path}''")
            
    def view(self):
        file_path = str(input("\nplease enter the file's path and name\nexample-->> C:\\Users\\PC\\Desktop\\sample.xlsx \n : "))

        wb = load_workbook(file_path)
        ws = wb.active
        
        q = str(input("\n1.show sheet names 2.show data from a sheet \n : "))
        
        if q=="1":
            
            all_sheet_names = wb.sheetnames
            print(f"sheet names : {all_sheet_names}")
        
        if q=="2":
        
            cell =  str(input("\nplease enter the cell \nexample-->> B5\n: \n"))
            
            print(ws[cell].value)

run=excel1()


while True:
    q="y"
    while q=="y":
        ask=str(input("\n📚 OPENPYXL 📚\n1.Create 2.Edit 3.View 4.Exit :  "))

        if ask=="1":

            run.create()

        if ask=="2":

            run.edit()

        if ask=="3":

            run.view()

        if ask=="4":
            break

        else:
            continue  

    q2=input("\nType 'q' to exit or anything else to continue: ").lower()

    if q2=="q":
        break

print("\nヾ(￣▽￣) Bye~Bye~")